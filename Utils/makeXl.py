from os import sep
from sys import argv
from os.path import join
from datetime import datetime
from sys import exit as sysexit
from warnings import filterwarnings
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from get_techstack_vulnerabilities import FetchTechStackVulnerabilities

class CreateXl:

    def __init__(self, xls_file_name=None, df_cvc=None, df_techstack=None):
        
        self.df_cvc = df_cvc
        self.df_techstack = df_techstack
        self.today = datetime.today().strftime('%d-%m-%Y')
        self.xls_file_name = join(*xls_file_name, 'dependency-check-report-' + str(self.today) + '.xlsx')
    
    def create(self):
        try:

            workbook = Workbook()
            workbook.remove(workbook.active)
            header_font = Font(name='Fira Code',bold=True,color='FFFFFF')
            centered_alignement = Alignment(horizontal='center', vertical='center')
            wrapped_alignement = Alignment(vertical='top', wrap_text=False)
            fill = PatternFill(start_color='5FABE6', end_color='5FABE6', fill_type='solid')
            
            if self.df_cvc is not None:
                cvc_sheet_columns = [
                    ('DependencyName', 40),
                    ('Description', 40),
                    ('CVE', 30),
                    ('Severity', 15),
                    ('FilePath', 40),
                    ('Status', 15),
                    ('Auditor Comment', 40),
                    ('Developer Comment', 40)
                ]

                worksheet = workbook.create_sheet(title='CVC', index=0)
                row_num = 1
                for col_num, (column_title, column_width) in enumerate(cvc_sheet_columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title
                    cell.font = header_font
                    cell.alignment = centered_alignement
                    cell.fill = fill
                    column_letter = get_column_letter(col_num)
                    column_dimensions = worksheet.column_dimensions[column_letter]
                    column_dimensions.width = column_width
                
                for i in range(len(self.df_cvc)):
                    row_num+=1
                    row = [
                        (self.df_cvc.loc[i, 'DependencyName'], 'Normal'),
                        (self.df_cvc.loc[i, 'Description'], 'Normal'),
                        (self.df_cvc.loc[i, 'CVE'], 'Normal'),
                        (self.df_cvc.loc[i, 'Severity'], 'Normal'),
                        (self.df_cvc.loc[i, 'FilePath'], 'Normal'),
                        (self.df_cvc.loc[i, 'Status'], 'Normal'),
                        (self.df_cvc.loc[i, 'Auditor Comment'], 'Normal'),
                    ]

                    for col_num, (cell_value, cell_format) in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.style = cell_format
                        cell.alignement = wrapped_alignement

            if self.df_techstack is not None:
                tech_stack_sheet_columns = [
                    ('DependencyName', 40),
                    ('Descrption', 40),
                    ('CVE', 30),
                    ('Severity', 15),
                    ('Status', 15),
                    ('Auditor Comments', 40),
                    ('Developer Comments', 40)
                ]

                index_value = 1 if self.df_cvc is not None else 0
                worksheet = workbook.create_sheet(title='TechStack', index=index_value)
                row_num = 1
                for col_num, (column_title, column_width) in enumerate(tech_stack_sheet_columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title
                    cell.font = header_font
                    cell.fill = fill
                    cell.alignment = centered_alignement
                    column_letter = get_column_letter(col_num)
                    column_dimensions = worksheet.column_dimensions[column_letter]
                    column_dimensions.width = column_width
                for i in range(len(self.df_techstack)):
                    row_num+=1
                    row = [
                        (self.df_techstack.loc[i, 'Product'], 'Normal'),
                        (self.df_techstack.loc[i, 'Description'], 'Normal'),
                        (self.df_techstack.loc[i, 'CVE'], 'Normal'),
                        (self.df_techstack.loc[i, 'Severity'], 'Normal'),
                        (self.df_techstack.loc[i, 'Status'], 'Normal'),
                        (self.df_techstack.loc[i, 'Auditor Comment'], 'Normal'),
                    ]

                    for col_num, (cell_value, cell_format) in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.style = cell_format
                    
                worksheet.freeze_panes = worksheet['A2']
                worksheet.sheet_properties.tabColor = '5FABE6'
            workbook.save(self.xls_file_name)
            print("[Info] Excel created successfully....")
        except Exception as e:
            print("Unable to create xls....")
            sysexit(e)

if __name__ == "__main__":

    filterwarnings('ignore')

    techstack = FetchTechStackVulnerabilities()
    df_tech_stack = techstack.techStackDataToDf()
    output_location = argv[1].split("\\")
    output_location.insert(1, sep)
    
    xls = CreateXl(
        xls_file_name=output_location,
        df_cvc=None,
        df_techstack=df_tech_stack
    )
    xls.create()